import pygame
import os
from openpyxl import Workbook, load_workbook
from audiofile_downloader import RunScrapper


pygame.mixer.init()



audio_dir = "downloaded_files"
excel_file = "missing_words.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Words"
    ws.append(["Word"])
    wb.save(excel_file)

wb = load_workbook(excel_file)
ws = wb["Missing Words"]


saved_word_list = []
for row in ws.iter_rows(min_row=2, max_col=1):
    saved_word_list.append(row[0].value)



class TTS:
    def RunTTS(self):
        sentence = input("Enter text: ")
        text = sentence.split()

        missing_words = set()

        for word in text:
            audio_file = f"{audio_dir}/{word}.mp3"

            if not os.path.exists(audio_file):
                missing_words.add(word)



        if len(missing_words)==0:
            for word in text:
                audio_file = f"{audio_dir}/{word}.mp3"
                pygame.mixer.music.load(audio_file)
                pygame.mixer.music.play()
                while pygame.mixer.music.get_busy():
                    pygame.time.Clock().tick(10)                
                
        elif len(missing_words)>0:
            print(f"Collecting some missing data. Please wait a moment")
            collector=RunScrapper()
            collector.download_audios(missing_words)
            print("Download Successful")
            for word in text:
                audio_file = f"{audio_dir}/{word}.mp3"
                pygame.mixer.music.load(audio_file)
                pygame.mixer.music.play()
                while pygame.mixer.music.get_busy():
                    pygame.time.Clock().tick(10)                  




        # for word in missing_words:
        #     word_exists = False
        #     for row in ws.iter_rows(min_row=2, max_col=1):
        #         if word == row[0].value:
        #             word_exists = True
        #             break
        #     if not word_exists:
        #         ws.append([word])

        # wb.save(excel_file)
        # print(f"Missing words have been saved to {excel_file}.")

        

        # if len(saved_word_list)>=2:
        #     print(f"{len(saved_word_list)} words as been saved to {excel_file}.")
        #     user_choice=input("Download voice for all of them(y/n)? : ")
        #     if user_choice=="y":
        #         print("Downloading missing files.....")

        #         downloader=RunScrapper()
        #         downloader.download_audios(saved_word_list)
            
        #         ws.delete_rows(2,ws.max_row)
        #         wb.save(excel_file)
        #         print(f"All words have been downloaded to the {audio_dir} and deleted from the {excel_file}.")
        #     elif user_choice == "n":
        #         print("Please download all the files next time")
        #     else:
        #         print("Wrong Input")


# print("Press 1 to run TTS.")
# print("Press 2 to download the missing files")
# choice=int(input())
# if choice == 1:
#     run=TTS()
#     run.RunTTS()
# elif choice == 2:
#     if len(saved_word_list)==0:
#         print("The missing word list is empty now")
#     else:
#         print("Downloading missing files.....")
#         downloader=RunScrapper()
#         downloader.download_audios(saved_word_list)
#         ws.delete_rows(2,ws.max_row)
#         wb.save(excel_file)
#         print(f"All words have been downloaded to the {audio_dir} and deleted from the {excel_file}.")    

# else :
#     print("Wrong Input")


run=TTS()
run.RunTTS()